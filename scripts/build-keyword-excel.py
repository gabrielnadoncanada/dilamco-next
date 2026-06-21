# -*- coding: utf-8 -*-
"""
Construit un Excel d'analyse de mots-clés (Google Keyword Planner) pour structurer
la boutique en ligne Dilamco (vanités, armoires, comptoirs, liquidation, etc.).
Source : export CSV Keyword Planner (UTF-16, tab). Marché : Canada / français.
"""
import csv, re, unicodedata, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRCS = [
    r"C:\Users\gabri\OneDrive\Pictures\Keyword Stats 2026-06-16 at 14_54_38.csv",
    r"C:\Users\gabri\OneDrive\Pictures\Keyword Stats 2026-06-16 at 15_30_06.csv",
]
OUT = r"C:\laragon\www\dilamco-next\dilamco-recherche-mots-cles-boutique.xlsx"

# ---------- lecture + fusion + déduplication (on garde le volume le plus élevé) ----------
period = ""
merged = {}  # clé = mot-clé en minuscules -> ligne brute
for SRC in SRCS:
    raw = open(SRC, "rb").read().decode("utf-16")
    lines = raw.splitlines()
    if not period:
        period = lines[1].strip().strip('"')
    for l in lines[3:]:
        if not l.strip():
            continue
        cells = l.split("\t")
        kw = cells[0].strip()
        if not kw:
            continue
        key = kw.lower()
        def _v(c):
            try: return int(c[2])
            except: return -1
        if key not in merged or _v(cells) > _v(merged[key]):
            merged[key] = cells
data = list(merged.values())

def to_num(v):
    try:
        return int(v)
    except:
        return None

VOL_LABEL = {50000:"10 K – 100 K", 5000:"1 K – 10 K", 500:"100 – 1 K",
             50:"10 – 100", 0:"0 – 10", None:"n/d"}
COMP_FR = {"High":"Élevée", "Medium":"Moyenne", "Low":"Faible", "Unknown":"n/d", "":"n/d"}

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

# ---------- dictionnaires de classification ----------
MATERIAUX = ["quartz","stratifie","dekton","granit","melamine","marbre","bois","noyer",
             "chene","thermoplastique","mdf","ceramique","porcelaine","acrylique","inox",
             "beton","corian","laque","verre"]
COULEURS  = ["blanc","blanche","noir","noire","gris","grise","beige","brun","brune","bleu",
             "bleue","vert","verte","taupe","anthracite","naturel","creme","ivoire"]
MARQUES   = ["ikea","home depot","homedepot","rona","canac","costco","reno depot","reno-depot",
             "tanguay","mobilia","wayfair","leon","brault","patio","cuisines action","kitchen"]
PRIX      = ["liquidation","solde","rabais","pas cher","prix","clearance","cheap","discount",
             "aubaine","escompte","bon marche","abordable","economique","fin de serie","surplus"]
INFO      = ["comment","installer","installation","dimension","hauteur","largeur","profondeur",
             "standard","idee","diy","fabriquer","peinturer","peindre","nettoyer","entretien",
             "mesure","plan","reglementaire","norme"]

DIM_RE = re.compile(r"(\d{2,3})\s*(pouce|pouces|po|cm|\"|pi|pied|pieds|x)")

def detect_list(t, words):
    found = [w for w in words if w in t]
    return found

def categorie(t):
    # ordre important : spécifique -> général
    if "pharmacie" in t or "armoire a pharmacie" in t or ("miroir" in t and "salle" in t):
        return "Pharmacie & miroir (SdB)"
    if "vanite" in t or "vanity" in t:
        return "Vanité / meuble-lavabo (SdB)"
    if "meuble" in t and ("lavabo" in t or "vasque" in t or ("salle" in t and "bain" in t)):
        return "Vanité / meuble-lavabo (SdB)"
    if "vasque" in t or "lavabo" in t or ("evier" in t and "salle" in t) or ("evier" in t and "bain" in t):
        return "Lavabo / vasque / évier (SdB)"
    if "comptoir" in t or "plan de travail" in t or "dosseret" in t or "ilot" in t or "dekton" in t and "comptoir" in t:
        return "Comptoir & surface"
    if "comptoir" in t:
        return "Comptoir & surface"
    if ("armoire" in t or "armoir" in t or "caisson" in t or "porte" in t or "facade" in t) and ("cuisine" in t):
        return "Armoire de cuisine"
    if "caisson" in t or ("porte" in t and ("armoire" in t or "cuisine" in t)) or "facade" in t:
        return "Armoire de cuisine"
    if "armoire" in t and ("salle" in t or "bain" in t):
        return "Rangement / armoire (SdB)"
    if "rangement" in t or "etagere" in t or "colonne" in t or "tablette" in t or "tour de rangement" in t:
        return "Rangement / étagère (SdB)"
    if "poignee" in t or "charniere" in t or "coulisse" in t or "tiroir" in t or "quincaillerie" in t:
        return "Quincaillerie"
    if "robinet" in t or "douche" in t or "baignoire" in t or "toilette" in t or "bain " in t:
        return "Plomberie & accessoires"
    if "armoire" in t or "cuisine" in t:
        return "Armoire de cuisine"
    return "Autre / général"

def piece(t):
    sdb = any(k in t for k in ["salle de bain","salle bain","salle d bain","sdb","vanite","vanity","lavabo","vasque","pharmacie"])
    cui = "cuisine" in t
    if sdb and cui: return "Mixte"
    if sdb: return "Salle de bain"
    if cui: return "Cuisine"
    return "Général"

def intention(t):
    if any(k in t for k in PRIX): return "Transactionnel (prix/liquidation)"
    if any(k in t for k in MARQUES): return "Marque / concurrent"
    if DIM_RE.search(t) or any(m in t for m in MATERIAUX) or any(c in t for c in COULEURS):
        return "Commercial (attribut)"
    if any(k in t for k in INFO): return "Informationnel"
    return "Générique produit"

# ---------- transformation ----------
records = []
for r in data:
    kw = r[0].strip()
    if not kw: continue
    t = strip_accents(kw.lower())
    vol = to_num(r[2])
    comp_idx = to_num(r[6]) if len(r) > 6 else None
    bid_low = r[7] if len(r) > 7 and r[7] not in ("","0") else ""
    bid_high = r[8] if len(r) > 8 and r[8] not in ("","0") else ""
    mat = detect_list(t, MATERIAUX)
    col = [c for c in COULEURS if re.search(r"\b"+c+r"\b", t)]
    dim = DIM_RE.search(t)
    mar = [m for m in MARQUES if m in t]
    records.append({
        "kw": kw,
        "vol": vol if vol is not None else -1,
        "vol_label": VOL_LABEL.get(vol, "n/d"),
        "comp": COMP_FR.get(r[5], "n/d"),
        "comp_idx": comp_idx if comp_idx is not None else "",
        "bid_low": bid_low,
        "bid_high": bid_high,
        "piece": piece(t),
        "cat": categorie(t),
        "intent": intention(t),
        "materiau": ", ".join(mat),
        "couleur": ", ".join(sorted(set(col))),
        "dimension": dim.group(0) if dim else "",
        "marque": ", ".join(mar),
    })

records.sort(key=lambda x: (-x["vol"], x["kw"]))

# ---------- styles ----------
HDR_FILL = PatternFill("solid", fgColor="253B2F")  # vert atelier Dilamco
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=15, color="253B2F")
SUB_FONT = Font(italic=True, size=10, color="6C7068")
THIN = Side(style="thin", color="DCD5C5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F6F3EC")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

wb = Workbook()

# ======== Feuille 1 : Lisez-moi ========
ws = wb.active
ws.title = "Lisez-moi"
ws["A1"] = "Dilamco — Recherche de mots-clés Google (boutique en ligne)"; ws["A1"].font = TITLE_FONT
ws["A2"] = f"Source : Google Keyword Planner · Marché : Canada · Langue : français · Période : {period}"
ws["A2"].font = SUB_FONT
readme = [
 "",
 "OBJECTIF",
 "Structurer la boutique (catégories, sous-catégories, fiches produits, attributs/filtres, descriptions, fiches techniques)",
 "pour capter un maximum de recherches Google, de trafic et de ventes.",
 "",
 "CE QUE CONTIENT CE FICHIER",
 "• « Tous les mots-clés » : 8 620 mots-clés, volume mensuel, concurrence, enchère, + classification (pièce, catégorie, intention, attributs). Filtrable.",
 "• « Synthèse catégories » : nb de mots-clés et volume estimé par catégorie de produit → où est la demande.",
 "• « Structure boutique » : arborescence recommandée (catégories → sous-catégories → filtres) mappée à la demande réelle.",
 "• « Top opportunités » : mots-clés à fort volume + cas liquidation et faible concurrence à prioriser.",
 "",
 "LECTURE DES VOLUMES (limite des comptes Google Ads sans campagne active)",
 "Google fournit des TRANCHES, pas un chiffre exact. On utilise la valeur représentative de chaque tranche :",
 "   10 K–100 K → 50 000   |   1 K–10 K → 5 000   |   100–1 K → 500   |   10–100 → 50   |   0–10 → 0",
 "Le « Volume estimé » sommé sert à COMPARER des groupes entre eux, pas comme prévision absolue.",
 "",
 "CONCURRENCE = concurrence PUBLICITAIRE (Google Ads), pas la difficulté SEO. Utile comme indice de valeur commerciale.",
 "« Indice concurrence » (0–100) affine ce signal. L'enchère (CA$) indique la valeur d'un clic payant = intention d'achat.",
 "",
 "MÉTHODE DE PRIORISATION SUGGÉRÉE",
 "1) Fort volume + intention transactionnelle → catégories et pages piliers.",
 "2) Longue traîne (attributs : matériau, couleur, dimension) → filtres de boutique + variantes de fiches produits.",
 "3) « Liquidation / prix » → pages et collections dédiées (vendre vite le stock).",
]
row = 3
for line in readme:
    ws.cell(row=row, column=1, value=line)
    if line.isupper() and line.strip():
        ws.cell(row=row, column=1).font = Font(bold=True, color="253B2F", size=11)
    row += 1
ws.column_dimensions["A"].width = 130

# ======== Feuille 2 : Tous les mots-clés ========
ws = wb.create_sheet("Tous les mots-clés")
cols = ["Mot-clé","Volume mensuel","Volume estimé","Concurrence","Indice conc. (0-100)",
        "Enchère bas (CA$)","Enchère haut (CA$)","Pièce","Catégorie produit","Intention",
        "Matériau","Couleur","Dimension","Marque/concurrent"]
ws.append(cols)
style_header(ws, 1, len(cols))
for i, r in enumerate(records):
    ws.append([
        r["kw"], r["vol_label"], (r["vol"] if r["vol"]>=0 else ""), r["comp"], r["comp_idx"],
        r["bid_low"], r["bid_high"], r["piece"], r["cat"], r["intent"],
        r["materiau"], r["couleur"], r["dimension"], r["marque"],
    ])
    if i % 2 == 1:
        for c in range(1, len(cols)+1):
            ws.cell(row=i+2, column=c).fill = ALT_FILL
widths = [40,15,13,12,14,14,15,15,28,30,18,16,14,20]
for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(records)+1}"

# ======== Feuille 3 : Synthèse catégories ========
ws = wb.create_sheet("Synthèse catégories")
agg = defaultdict(lambda: {"n":0,"vol":0,"high":0})
for r in records:
    a = agg[r["cat"]]; a["n"]+=1
    if r["vol"]>0: a["vol"]+=r["vol"]
    if r["comp"]=="Élevée": a["high"]+=1
ws["A1"] = "Synthèse par catégorie de produit"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Trié par volume estimé décroissant — montre où se concentre la demande."; ws["A2"].font = SUB_FONT
hdr = ["Catégorie produit","Nb de mots-clés","Volume estimé / mois","% concurrence élevée"]
ws.append([]); ws.append(hdr); style_header(ws, 4, len(hdr))
for cat,a in sorted(agg.items(), key=lambda kv:-kv[1]["vol"]):
    pct = round(100*a["high"]/a["n"]) if a["n"] else 0
    ws.append([cat, a["n"], a["vol"], f"{pct}%"])
for i,w in enumerate([30,18,22,22]): ws.column_dimensions[get_column_letter(i+1)].width = w

# piece breakdown
start = ws.max_row + 3
ws.cell(row=start, column=1, value="Synthèse par pièce").font = Font(bold=True, size=12, color="253B2F")
ws.append([]) ; hdr2=["Pièce","Nb de mots-clés","Volume estimé / mois"]
ws.append(hdr2); style_header(ws, start+2, len(hdr2))
pag = defaultdict(lambda:{"n":0,"vol":0})
for r in records:
    p=pag[r["piece"]]; p["n"]+=1
    if r["vol"]>0: p["vol"]+=r["vol"]
for p,a in sorted(pag.items(), key=lambda kv:-kv[1]["vol"]):
    ws.append([p,a["n"],a["vol"]])

# intent breakdown
start2 = ws.max_row + 3
ws.cell(row=start2, column=1, value="Synthèse par intention de recherche").font = Font(bold=True, size=12, color="253B2F")
ws.append([]); hdr3=["Intention","Nb de mots-clés","Volume estimé / mois"]
ws.append(hdr3); style_header(ws, start2+2, len(hdr3))
iag = defaultdict(lambda:{"n":0,"vol":0})
for r in records:
    a=iag[r["intent"]]; a["n"]+=1
    if r["vol"]>0: a["vol"]+=r["vol"]
for k,a in sorted(iag.items(), key=lambda kv:-kv[1]["vol"]):
    ws.append([k,a["n"],a["vol"]])

# ======== Feuille 4 : Structure boutique recommandée ========
ws = wb.create_sheet("Structure boutique")
ws["A1"] = "Arborescence de boutique recommandée (basée sur la demande réelle)"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Catégorie → sous-catégories → filtres/attributs. À utiliser pour l'URL, le menu, les pages collection et les filtres."
ws["A2"].font = SUB_FONT
hdr = ["Niveau 1 (catégorie)","Niveau 2 (sous-catégories / collections)","Filtres & attributs (facettes)","Mots-clés cibles principaux","Note SEO / contenu"]
ws.append([]); ws.append(hdr); style_header(ws, 4, len(hdr))
structure = [
 ["Vanités & meubles-lavabos",
  "Vanité 1 robinet · Vanité double · Meuble sur pied · Meuble suspendu · Avec / sans comptoir · Avec miroir",
  "Largeur (24/30/36/48/60 po) · Couleur · Matériau · Type de vasque · Nb de tiroirs",
  "vanité salle de bain, meuble lavabo, meuble-vasque, vanité 48 pouces, vanité double",
  "Catégorie PILIER (volume le plus élevé). Page collection riche + guide des tailles + fiches par largeur."],
 ["Lavabos, vasques & éviers",
  "Vasque à poser · Vasque encastrée · Lavabo sur colonne · Évier de salle de bain",
  "Forme · Matériau (céramique/porcelaine) · Couleur · Mode d'installation",
  "vasque salle de bain, lavabo salle de bain, évier salle de bain",
  "Souvent acheté avec la vanité → cross-sell + bundles."],
 ["Pharmacies & miroirs",
  "Pharmacie encastrée · Pharmacie en surface · Miroir éclairé · Miroir avec rangement",
  "Largeur · Éclairage LED (oui/non) · Anti-buée · Couleur",
  "pharmacie salle de bain, armoire à pharmacie, miroir salle de bain",
  "Fort volume distinct. Page dédiée + lien depuis chaque vanité."],
 ["Armoires de cuisine",
  "Armoires complètes · Caissons · Portes & façades · Armoires en stock (prêtes)",
  "Matériau (mélamine/thermoplastique/bois) · Couleur · Style (shaker/moderne) · Configuration",
  "armoire de cuisine, armoire cuisine, caisson de cuisine, porte d'armoire",
  "Catégorie cœur de métier. Mettre en avant « en stock » et délais."],
 ["Comptoirs & surfaces",
  "Quartz · Stratifié · Dekton · Granit · Bois · Comptoir de salle de bain",
  "Matériau · Couleur/fini · Épaisseur · Pièce (cuisine/SdB) · Sur mesure",
  "comptoir quartz, comptoir stratifié, dekton comptoir, comptoir de cuisine",
  "Beaucoup de variantes par matériau → 1 page par matériau (forte demande)."],
 ["Rangement & étagères (SdB)",
  "Colonnes · Étagères · Tours de rangement · Meubles d'appoint",
  "Hauteur · Couleur · Matériau · Nb de tablettes",
  "meuble rangement salle de bain, étagère salle de bain, colonne salle de bain",
  "Longue traîne utile pour fiches + filtres."],
 ["Quincaillerie",
  "Poignées · Charnières · Coulisses · Accessoires de tiroir",
  "Fini · Type · Format",
  "poignée armoire, charnière armoire, coulisse tiroir",
  "Panier moyen faible mais récurrent / accessoire."],
 ["Liquidation & aubaines",
  "Liquidation vanités · Liquidation armoires de cuisine · Fin de série · Surplus",
  "Type de produit · Rabais (%) · Disponibilité",
  "liquidation armoire de cuisine, vanité salle de bain liquidation, liquidation vanité",
  "COLLECTION TRANSVERSALE prioritaire : vendre le stock vite. Page + badges « Liquidation » sur fiches."],
]
r0 = 5
for i,rowv in enumerate(structure):
    ws.append(rowv)
    for c in range(1, len(hdr)+1):
        cell = ws.cell(row=r0+i, column=c)
        cell.alignment = WRAP; cell.border = BORDER
        if i % 2 == 1: cell.fill = ALT_FILL
for i,w in enumerate([26,46,42,42,46]): ws.column_dimensions[get_column_letter(i+1)].width = w

# bloc conseils fiches produits / specs
b = ws.max_row + 3
tips = [
 ("MODÈLE DE FICHE PRODUIT (pour le SEO et la conversion)", 13),
 ("• Titre = [Type] + [attribut clé] + [dimension] : ex. « Vanité de salle de bain 48 po, 2 tiroirs, blanc mat ».", 11),
 ("• URL = /boutique/vanites/vanite-48-pouces-blanc (mots-clés, pas d'ID).", 11),
 ("• Description : 1er paragraphe répond à l'intention (usage, pièce, bénéfice) + mots-clés naturels.", 11),
 ("• Fiche technique (spec sheet) : dimensions exactes, matériau, fini, type d'installation, inclus/non inclus, garantie, entretien.", 11),
 ("• Données structurées Product + Offer + AggregateRating + Breadcrumb sur chaque fiche.", 11),
 ("• Variantes (couleur/dimension) regroupées sur une même fiche quand pertinent (évite la cannibalisation).", 11),
 ("• Bilingue : chaque page a son équivalent /en avec hreflang (cohérent avec le site vitrine).", 11),
]
for j,(txt,sz) in enumerate(tips):
    cell = ws.cell(row=b+j, column=1, value=txt)
    cell.font = Font(bold=(sz==13), size=sz, color="253B2F" if sz==13 else "1A1F1C")

# ======== Feuille 5 : Top opportunités ========
ws = wb.create_sheet("Top opportunités")
ws["A1"] = "Top opportunités à prioriser"; ws["A1"].font = TITLE_FONT
ws["A2"] = "A) Plus gros volumes  ·  B) Liquidation/prix (vendre le stock)  ·  C) Volume avec concurrence pub plus faible."
ws["A2"].font = SUB_FONT

def block(title, rows, startrow):
    ws.cell(row=startrow, column=1, value=title).font = Font(bold=True, size=12, color="253B2F")
    hdr = ["Mot-clé","Volume mensuel","Volume estimé","Concurrence","Catégorie produit","Intention"]
    for c,h in enumerate(hdr,1):
        ws.cell(row=startrow+1, column=c)
    ws.append([]) if False else None
    for c,h in enumerate(hdr,1):
        cell = ws.cell(row=startrow+1, column=c, value=h)
    style_header(ws, startrow+1, len(hdr))
    rr = startrow+2
    for r in rows:
        ws.cell(row=rr, column=1, value=r["kw"])
        ws.cell(row=rr, column=2, value=r["vol_label"])
        ws.cell(row=rr, column=3, value=(r["vol"] if r["vol"]>=0 else ""))
        ws.cell(row=rr, column=4, value=r["comp"])
        ws.cell(row=rr, column=5, value=r["cat"])
        ws.cell(row=rr, column=6, value=r["intent"])
        rr += 1
    return rr

topA = [r for r in records if r["vol"]>=500][:60]
r1 = block("A) Plus gros volumes (≥ 100 rech./mois)", topA, 4)
liq = [r for r in records if r["intent"]=="Transactionnel (prix/liquidation)" and r["vol"]>=50]
liq.sort(key=lambda x:-x["vol"])
r2 = block("B) Liquidation / prix — vendre le stock vite", liq[:40], r1+2)
opp = [r for r in records if r["vol"]>=500 and r["comp"] in ("Faible","Moyenne")]
opp.sort(key=lambda x:-x["vol"])
r3 = block("C) Volume avec concurrence publicitaire plus faible", opp[:40], r2+2)
for i,w in enumerate([42,16,14,13,28,30]): ws.column_dimensions[get_column_letter(i+1)].width = w

wb.save(OUT)
print("OK ->", OUT)
print("records:", len(records))
print("cats:", dict((c,agg[c]["n"]) for c in agg))
